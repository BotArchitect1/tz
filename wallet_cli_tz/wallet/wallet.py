import logging
from datetime import date, datetime
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from wallet_cli_tz.transaction.transaction import Transaction


class Wallet:
    """
    Класс для управления списком транзакций и реализации основных операций.

    Attributes:
        transactions (List[Transaction]): Список всех транзакций.
        file_path (str): Путь к файлу Exel для хранения данных.
    """

    def __init__(self, file_path: str):
        self.transactions = []
        self.file_path = file_path
        self.load_transactions()

    def get_balance(self) -> Tuple[float, float, float]:
        total_income = sum(
            tx.amount for tx in self.transactions if tx.category == "Доход"
        )
        total_expense = sum(
            tx.amount for tx in self.transactions if tx.category == "Расход"
        )
        balance = total_income - total_expense
        return balance, total_income, total_expense

    def add_transaction(
        self,
        date: date,
        category: str,
        amount: float,
        description: Optional[str] = None,
    ):

        transaction = Transaction(date, category, amount, description)
        self.transactions.append(transaction)
        self.save_transactions()

    def edit_transaction(self, index: int, **kwargs):
        transaction = self.transactions[index]
        for key, value in kwargs.items():
            if hasattr(transaction, key):
                setattr(transaction, key, value)
        self.save_transactions()

    def search_transactions(
        self,
        category: Optional[str] = None,
        date: Optional[date] = None,
        amount: Optional[float] = None,
    ) -> List[Transaction]:
        result = []
        for transaction in self.transactions:
            if (
                (category is None or transaction.category == category)
                and (date is None or transaction.date == date)
                and (amount is None or transaction.amount == amount)
            ):
                result.append(transaction)
        return result

    def load_transactions(self):
        try:
            wb = load_workbook(filename=self.file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                date_transaction = datetime.strptime(row[0], "%Y-%m-%d").date()
                category = row[1]
                amount = float(row[2])
                description = row[3]
                transaction = Transaction(
                    date_transaction, category, amount, description
                )
                self.transactions.append(transaction)
            wb.close()
        except FileNotFoundError:
            logging.info("File not found")

    def save_transactions(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Category", "Amount", "Description"])
        for transaction in self.transactions:
            ws.append(
                [
                    transaction.date.strftime("%Y-%m-%d"),
                    transaction.category,
                    transaction.amount,
                    transaction.description,
                ]
            )
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    logging.info(f"An error {e} occurred")
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
            ws.column_dimensions[column].alignment = Alignment(wrap_text=True)
        wb.save(filename=self.file_path)
